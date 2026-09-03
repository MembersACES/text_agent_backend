"""Fill the DMA contract-details workbook from the on-disk template (no Drive)."""

from io import BytesIO
from unittest.mock import MagicMock

from openpyxl import load_workbook

from tools.dma_contract_details import (
    TEMPLATE_PATH,
    build_filename,
    fill_workbook_bytes,
    resolve_target_folder_id,
    row_values,
)

PAYLOAD = {
    "nmi": "61020000000",
    "business": "RSL Victoria",
    "abn": "12 345 678 901",
    "postal_address": "4 Collins St, Melbourne VIC 3000",
    "site_address": "1 Anzac Ave, Frankston VIC 3199",
    "frmp": "Alinta Energy",
    "contact": "Clint Example",
    "position": "CEO",
    "telephone": "03 9000 0000",
    "email": "clint@example.com",
    "dma_price": "600",
    "vas_price": "300",
    "start_date": "2026-09-01",
    "end_date": "2031-09-01",
}


def test_template_exists():
    assert TEMPLATE_PATH.is_file()


def test_row_values_formats_dates_and_numbers():
    values = row_values(PAYLOAD)
    assert values["start_date"] == "01/09/2026"
    assert values["end_date"] == "01/09/2031"
    assert values["meter"] == 600
    assert values["vas"] == 300
    assert values["nmi"] == "61020000000"


def test_fill_workbook_maps_typo_headers():
    raw = fill_workbook_bytes(PAYLOAD)

    wb = load_workbook(BytesIO(raw))
    ws = wb.active
    headers = [str(cell.value or "").strip() for cell in ws[1]]
    filled = {headers[i]: ws.cell(row=2, column=i + 1).value for i in range(len(headers)) if headers[i]}
    assert filled["NMI"] == "61020000000"
    assert filled["Business"] == "RSL Victoria"
    assert filled["ABN"] == "12 345 678 901"
    assert filled["FRMP"] == "Alinta Energy"
    assert filled["Postion"] == "CEO"
    assert filled["Meter"] == 600
    assert filled["VAS"] == 300
    assert filled["DMA Start date"] == "01/09/2026"
    assert filled["Dma END DATE"] == "01/09/2031"


def test_build_filename_strips_unsafe_chars():
    name = build_filename("RSL Victoria / Frankston", "6102/000")
    assert name.startswith("DMA contract details - ")
    assert name.endswith(".xlsx")
    assert "/" not in name


def test_resolve_target_folder_prefers_engagement_form_parent():
    drive = MagicMock()
    drive.files().get().execute.return_value = {
        "id": "file123",
        "mimeType": "application/vnd.google-apps.document",
        "parents": ["folderABC"],
    }
    folder_id = resolve_target_folder_id(
        engagement_form_link="https://drive.google.com/file/d/file123/view",
        client_folder_url="https://drive.google.com/drive/folders/fallbackXYZ",
        drive_service=drive,
    )
    assert folder_id == "folderABC"


def test_resolve_target_folder_falls_back_to_client_folder():
    folder_id = resolve_target_folder_id(
        engagement_form_link="",
        client_folder_url="https://drive.google.com/drive/folders/fallbackXYZ",
        drive_service=None,
    )
    assert folder_id == "fallbackXYZ"
