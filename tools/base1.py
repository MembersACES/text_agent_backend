"""
Base 1 Review - Quick Win Tool
Handles document upload, stub extraction, and output generation
"""
import os
import json
import uuid
import logging
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib import colors
import PyPDF2

logger = logging.getLogger(__name__)

# Storage base directory
STORAGE_BASE = Path("storage")
BASE1_STORAGE = STORAGE_BASE / "base1"
RUNS_FILE = BASE1_STORAGE / "runs.json"

# Ensure storage directories exist
BASE1_STORAGE.mkdir(parents=True, exist_ok=True)

# Maximum file size: 20MB
MAX_FILE_SIZE = 20 * 1024 * 1024


def init_storage():
    """Initialize storage directories"""
    BASE1_STORAGE.mkdir(parents=True, exist_ok=True)
    if not RUNS_FILE.exists():
        with open(RUNS_FILE, 'w') as f:
            json.dump({}, f)
    logger.info(f"Base1 storage initialized at {BASE1_STORAGE}")


def get_runs() -> Dict:
    """Load runs from JSON file"""
    if not RUNS_FILE.exists():
        return {}
    try:
        with open(RUNS_FILE, 'r') as f:
            return json.load(f)
    except Exception as e:
        logger.error(f"Error loading runs: {e}")
        return {}


def save_runs(runs: Dict):
    """Save runs to JSON file"""
    try:
        with open(RUNS_FILE, 'w') as f:
            json.dump(runs, f, indent=2, default=str)
    except Exception as e:
        logger.error(f"Error saving runs: {e}")
        raise


def create_run(email: Optional[str] = None, state: Optional[str] = None, business_name: Optional[str] = None) -> str:
    """Create a new Base1 run and return run_id"""
    run_id = str(uuid.uuid4())
    runs = get_runs()
    
    runs[run_id] = {
        "run_id": run_id,
        "business_name": business_name,  # Will be extracted from invoices
        "email": email,
        "state": state,
        "created_at": datetime.now().isoformat(),
        "documents": [],
        "accounts": {},  # Group documents by NMI/MRIN: {"NMI123": [doc_ids], "MRIN456": [doc_ids]}
        "extracted": False,
        "generated": False
    }
    
    save_runs(runs)
    
    # Create run directory
    run_dir = BASE1_STORAGE / run_id
    run_dir.mkdir(exist_ok=True)
    (run_dir / "documents").mkdir(exist_ok=True)
    (run_dir / "outputs").mkdir(exist_ok=True)
    
    logger.info(f"Created Base1 run: {run_id} (email: {email or 'not provided'})")
    return run_id


def save_document(run_id: str, filename: str, file_bytes: bytes) -> Dict:
    """Save uploaded document and return document record"""
    runs = get_runs()
    if run_id not in runs:
        raise ValueError(f"Run {run_id} not found")
    
    # Validate file size
    if len(file_bytes) > MAX_FILE_SIZE:
        raise ValueError(f"File {filename} exceeds maximum size of {MAX_FILE_SIZE / (1024*1024):.1f}MB")
    
    # Validate PDF
    if not filename.lower().endswith('.pdf'):
        raise ValueError(f"File {filename} is not a PDF")
    
    # Get page count
    page_count = 0
    try:
        from io import BytesIO
        pdf_reader = PyPDF2.PdfReader(BytesIO(file_bytes))
        page_count = len(pdf_reader.pages)
    except Exception as e:
        logger.warning(f"Could not read PDF pages for {filename}: {e}")
    
    # Save file
    run_dir = BASE1_STORAGE / run_id / "documents"
    safe_filename = "".join(c for c in filename if c.isalnum() or c in "._- ")
    file_path = run_dir / safe_filename
    
    # Handle duplicate filenames
    counter = 1
    original_path = file_path
    while file_path.exists():
        stem = original_path.stem
        suffix = original_path.suffix
        file_path = run_dir / f"{stem}_{counter}{suffix}"
        counter += 1
    
    with open(file_path, 'wb') as f:
        f.write(file_bytes)
    
    # Create document record
    doc_id = str(uuid.uuid4())
    doc_record = {
        "doc_id": doc_id,
        "filename": filename,
        "saved_filename": file_path.name,
        "size_bytes": len(file_bytes),
        "page_count": page_count,
        "created_at": datetime.now().isoformat(),
        "status": "uploaded",
        "extracted_fields": None
    }
    
    # Add to run
    runs[run_id]["documents"].append(doc_record)
    save_runs(runs)
    
    logger.info(f"Saved document {filename} for run {run_id}")
    return doc_record


def guess_utility_type(filename: str) -> str:
    """Best-effort guess of utility type from filename"""
    filename_lower = filename.lower()
    
    # Check for C&I or SME indicators
    is_ci = 'c&i' in filename_lower or 'ci' in filename_lower or 'commercial' in filename_lower or 'industrial' in filename_lower
    is_sme = 'sme' in filename_lower or 'small' in filename_lower
    
    if any(word in filename_lower for word in ['electricity', 'elec', 'power', 'nmi']):
        if is_ci:
            return "Electricity C&I"
        elif is_sme:
            return "Electricity SME"
        return "Electricity"
    elif any(word in filename_lower for word in ['gas', 'natural', 'lpg', 'mrin']):
        if is_ci:
            return "Gas C&I"
        elif is_sme:
            return "Gas SME"
        return "Gas"
    elif any(word in filename_lower for word in ['water', 'h2o']):
        return "Water"
    elif any(word in filename_lower for word in ['waste', 'rubbish', 'garbage', 'bin']):
        return "Waste"
    elif any(word in filename_lower for word in ['oil', 'cooking']):
        return "Oil"
    elif any(word in filename_lower for word in ['cleaning', 'clean']):
        return "Cleaning"
    elif any(word in filename_lower for word in ['dma', 'demand']):
        return "DMA"
    else:
        return "Unknown"


def extract_business_name_from_filename(filename: str) -> Optional[str]:
    """Try to extract business name from filename (best effort)"""
    import re
    
    # Remove .pdf extension
    name = filename.replace(".pdf", "").replace(".PDF", "")
    
    # Remove common invoice-related words (case insensitive)
    name = re.sub(r'\b(invoice|bill|statement|receipt)\b', '', name, flags=re.IGNORECASE)
    
    # Remove date patterns (DD_MM_YYYY-DD_MM_YYYY or YYYY-MM-DD)
    name = re.sub(r'\d{2}[_-]\d{2}[_-]\d{4}[_-]\d{2}[_-]\d{2}[_-]\d{4}', '', name)
    name = re.sub(r'\d{4}[-/]\d{2}[-/]\d{2}', '', name)
    name = re.sub(r'\d{2}[-/]\d{2}[-/]\d{4}', '', name)
    
    # Remove common prefixes like "for", "to", "from"
    name = re.sub(r'\b(for|to|from|of)\b', '', name, flags=re.IGNORECASE)
    
    # Remove utility type indicators if they're at the start
    name = re.sub(r'^(C&I|SME|CI|Commercial|Industrial)\s+', '', name, flags=re.IGNORECASE)
    name = re.sub(r'\s+(Electricity|Gas|Water|Waste|Oil)\s+', ' ', name, flags=re.IGNORECASE)
    
    # Remove NMI/MRIN patterns (like NDDD00GD30) - alphanumeric codes 8-12 chars
    name = re.sub(r'\b[A-Z0-9]{8,12}\b', '', name)  # Remove NMI/MRIN codes
    
    # Remove trailing dashes, underscores, and spaces
    name = re.sub(r'[-_]+$', '', name)
    name = name.strip(' -_')
    
    # If we have something meaningful (more than 3 chars), return it
    if len(name) > 3:
        return name
    return None


def extract_text_from_pdf(file_path: str) -> str:
    """Extract text content from PDF file"""
    try:
        with open(file_path, 'rb') as file:
            pdf_reader = PyPDF2.PdfReader(file)
            text = ""
            for page in pdf_reader.pages:
                text += page.extract_text() + "\n"
            return text
    except Exception as e:
        logger.warning(f"Could not extract text from PDF {file_path}: {e}")
        return ""


def extract_business_name_from_text(text: str, filename: str) -> Optional[str]:
    """Extract business name from PDF text content"""
    import re
    
    # Get first 2000 characters (business name is usually at the top)
    text_start = text[:2000]
    
    # Split into lines for better analysis
    lines = text_start.split('\n')
    
    # Known supplier names to exclude
    supplier_names = ['origin', 'agl', 'energyaustralia', 'alinta', 'simply', 'lumo', 
                      'red energy', 'dodo', 'momentum', 'click', 'powershop', 'diamond', 
                      'tango', '1st energy', 'essential energy']
    
    # Pattern 1: Look for company name at the very start (usually line 2-3 after supplier logo)
    # Format: "COMPANY NAME LTD" or "COMPANY NAME PTY LTD"
    for i, line in enumerate(lines[:10]):  # Check first 10 lines
        line_upper = line.upper().strip()
        
        # Skip empty lines, supplier names, and common invoice words
        if not line_upper or len(line_upper) < 5:
            continue
        if any(supplier in line_upper for supplier in supplier_names):
            continue
        if line_upper in ['ELECTRICITY', 'GAS', 'WATER', 'INVOICE', 'BILL', 'STATEMENT']:
            continue
        
        # Look for company name with LTD, PTY, etc.
        company_match = re.match(r'^([A-Z][A-Za-z0-9\s&.,()-]{5,60}?)\s+(?:LTD|PTY|PTY\.?|LIMITED|INC|INCORPORATED|LLC)', line_upper)
        if company_match:
            name = company_match.group(0).strip()  # Get full match including LTD
            # Verify it's not a supplier
            if not any(supplier in name.lower() for supplier in supplier_names):
                return name
        
        # Also check for company name without suffix (if it's substantial)
        if re.match(r'^[A-Z][A-Za-z0-9\s&.,()-]{8,50}$', line_upper):
            # Check if next line is an address (PO BOX, street, etc.)
            if i + 1 < len(lines):
                next_line = lines[i + 1].upper().strip()
                if 'PO BOX' in next_line or re.match(r'^\d+', next_line) or any(word in next_line for word in ['STREET', 'ROAD', 'AVENUE', 'LANE', 'DRIVE']):
                    return line.strip()
    
    # Pattern 2: Look for "Customer ABN" - the text before it is usually the business name
    customer_abn_pattern = r'([A-Z][A-Za-z0-9\s&.,()-]{5,60}?)\s+Customer\s+ABN'
    customer_abn_match = re.search(customer_abn_pattern, text_start, re.MULTILINE)
    if customer_abn_match:
        name = customer_abn_match.group(1).strip()
        # Exclude supplier names
        if not any(supplier in name.lower() for supplier in supplier_names):
            if len(name) > 5:
                return name
    
    # Pattern 2: Look for "Bill to:" or "Customer:" patterns (but exclude payment method text)
    patterns = [
        r'Bill\s+to[:\s]+([A-Z][A-Za-z0-9\s&.,()-]{5,50}?)(?:\n|PO\s+BOX|Address|ABN|ACN|\d{4})',
        r'Customer[:\s]+([A-Z][A-Za-z0-9\s&.,()-]{5,50}?)(?:\n|PO\s+BOX|Address|ABN|ACN|\d{4})',
        r'Account\s+Name[:\s]+([A-Z][A-Za-z0-9\s&.,()-]{5,50}?)(?:\n|PO\s+BOX|Address|ABN|ACN|\d{4})',
        r'Service\s+Address[:\s]+([A-Z][A-Za-z0-9\s&.,()-]{5,50}?)(?:\n|PO\s+BOX|Address|ABN|ACN|\d{4})',
    ]
    
    # Exclude common false positives from payment sections
    exclude_patterns = [
        r'American Express',
        r'Visa',
        r'Mastercard',
        r'Credit Card',
        r'Payment',
        r'BPAY',
        r'Direct Debit',
    ]
    
    for pattern in patterns:
        matches = re.finditer(pattern, text_start, re.IGNORECASE | re.MULTILINE)
        for match in matches:
            name = match.group(1).strip()
            
            # Skip if it matches exclusion patterns
            if any(re.search(exclude, name, re.IGNORECASE) for exclude in exclude_patterns):
                continue
            
            # Clean up
            name = re.sub(r'\s+(Pty|Ltd|Limited|Inc|Incorporated|LLC|ABN|ACN).*$', '', name, flags=re.IGNORECASE)
            name = name.strip()
            
            if len(name) > 5 and len(name) < 100:
                return name
    
    # Pattern 3: Look for company name followed by ABN (Australian Business Number)
    # Format: "COMPANY NAME ABN 12 345 678 901"
    abn_pattern = r'([A-Z][A-Za-z0-9\s&.,()-]{5,50}?)\s+ABN\s+\d{2}\s+\d{3}\s+\d{3}\s+\d{3}'
    abn_match = re.search(abn_pattern, text_start, re.MULTILINE)
    if abn_match:
        name = abn_match.group(1).strip()
        # Remove supplier names if they appear
        name = re.sub(r'^(ORIGIN|AGL|ENERGYAUSTRALIA|ALINTA|SIMPLY|LUMO|RED|DODO|MOMENTUM|CLICK|POWERSHOP|DIAMOND|TANGO|1ST)\s+', '', name, flags=re.IGNORECASE)
        if len(name) > 5:
            return name
    
    # Pattern 4: Look for "Customer ABN" pattern (the text before "Customer ABN")
    customer_abn_pattern = r'([A-Z][A-Za-z0-9\s&.,()-]{5,50}?)\s+Customer\s+ABN'
    customer_abn_match = re.search(customer_abn_pattern, text_start, re.MULTILINE)
    if customer_abn_match:
        name = customer_abn_match.group(1).strip()
        if len(name) > 5:
            return name
    
    # Fallback to filename extraction
    return extract_business_name_from_filename(filename)


def extract_supplier_from_text(text: str, filename: str) -> str:
    """Extract supplier/retailer name from PDF text"""
    import re
    
    # Common supplier indicators
    patterns = [
        r'(?:From|Supplier|Retailer|Energy\s+Retailer)[:\s]+([A-Z][A-Za-z0-9\s&.,-]+?)(?:\n|ABN|ACN|Address)',
        r'^([A-Z][A-Za-z0-9\s&.,-]{3,30})\s+(?:Energy|Power|Gas|Electricity|Utilities)',
    ]
    
    # Known Australian energy retailers
    known_retailers = [
        'AGL', 'Origin Energy', 'EnergyAustralia', 'Alinta Energy', 'Simply Energy',
        'Lumo Energy', 'Red Energy', 'Dodo Power & Gas', 'Momentum Energy',
        'Click Energy', 'Powershop', 'Diamond Energy', 'Tango Energy', '1st Energy'
    ]
    
    text_upper = text.upper()
    for retailer in known_retailers:
        if retailer.upper() in text_upper:
            return retailer
    
    for pattern in patterns:
        match = re.search(pattern, text, re.IGNORECASE | re.MULTILINE)
        if match:
            supplier = match.group(1).strip()
            if len(supplier) > 3 and len(supplier) < 100:
                return supplier
    
    return ""


def extract_invoice_date_from_text(text: str) -> str:
    """Extract invoice date from PDF text"""
    import re
    from datetime import datetime
    
    # Common date patterns
    date_patterns = [
        r'Invoice\s+Date[:\s]+(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})',
        r'Bill\s+Date[:\s]+(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})',
        r'Date[:\s]+(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})',
        r'(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})',  # Generic date
    ]
    
    for pattern in date_patterns:
        matches = re.findall(pattern, text)
        if matches:
            # Return the first date found (usually the invoice date)
            return matches[0]
    
    return ""


def extract_total_from_text(text: str) -> str:
    """Extract total amount from PDF text"""
    import re
    
    # Common total patterns
    patterns = [
        r'Total\s+(?:Inc|Including)\s+GST[:\s]+\$?([\d,]+\.?\d*)',
        r'Total\s+Amount[:\s]+\$?([\d,]+\.?\d*)',
        r'Amount\s+Due[:\s]+\$?([\d,]+\.?\d*)',
        r'Total[:\s]+\$?([\d,]+\.?\d*)',
    ]
    
    for pattern in patterns:
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            return match.group(1)
    
    return ""


def extract_stub_fields(doc_record: Dict) -> Dict:
    """Perform stub extraction from PDF content - returns extracted fields"""
    filename = doc_record["filename"]
    
    # Get the saved file path
    run_id = None
    runs = get_runs()
    for rid, run in runs.items():
        for doc in run.get("documents", []):
            if doc.get("doc_id") == doc_record.get("doc_id"):
                run_id = rid
                break
        if run_id:
            break
    
    if not run_id:
        logger.warning(f"Could not find run for document {doc_record.get('doc_id')}")
        # Fallback to filename-only extraction
        business_name = extract_business_name_from_filename(filename)
        utility_type = guess_utility_type(filename)
    else:
        # Read PDF content
        file_path = BASE1_STORAGE / run_id / "documents" / doc_record.get("saved_filename", filename)
        
        if file_path.exists():
            pdf_text = extract_text_from_pdf(str(file_path))
            
            if pdf_text:
                # Extract from PDF content
                business_name = extract_business_name_from_text(pdf_text, filename)
                supplier = extract_supplier_from_text(pdf_text, filename)
                invoice_date = extract_invoice_date_from_text(pdf_text)
                total_inc_gst = extract_total_from_text(pdf_text)
                utility_type = guess_utility_type(pdf_text + " " + filename)  # Use both text and filename
                
                confidence = 0.5 if business_name or supplier else 0.3
                flags = ["STUB_EXTRACTION", "PDF_TEXT_EXTRACTED"]
            else:
                # Fallback if text extraction failed
                business_name = extract_business_name_from_filename(filename)
                supplier = ""
                invoice_date = ""
                total_inc_gst = ""
                utility_type = guess_utility_type(filename)
                confidence = 0.3
                flags = ["STUB_EXTRACTION", "PDF_TEXT_EXTRACTION_FAILED"]
        else:
            # File not found, use filename only
            business_name = extract_business_name_from_filename(filename)
            supplier = ""
            invoice_date = ""
            total_inc_gst = ""
            utility_type = guess_utility_type(filename)
            confidence = 0.3
            flags = ["STUB_EXTRACTION", "FILE_NOT_FOUND"]
    
    extracted = {
        "filename": filename,
        "size_bytes": doc_record["size_bytes"],
        "page_count": doc_record.get("page_count", 0),
        "utility_type": utility_type,
        "business_name": business_name,
        "supplier": supplier,
        "site_address": "",
        "invoice_date": invoice_date,
        "total_inc_gst": total_inc_gst,
        "quantity": "",
        "units": "",
        "confidence": confidence,
        "flags": flags
    }
    
    return extracted


def run_extraction(run_id: str) -> List[Dict]:
    """Run stub extraction on all documents in a run"""
    runs = get_runs()
    if run_id not in runs:
        raise ValueError(f"Run {run_id} not found")
    
    run = runs[run_id]
    extracted_docs = []
    extracted_business_names = []
    
    for doc in run["documents"]:
        if doc.get("status") == "uploaded":
            extracted_fields = extract_stub_fields(doc)
            doc["extracted_fields"] = extracted_fields
            doc["status"] = "extracted"
            extracted_docs.append(doc)
            
            # Collect business names from extraction
            if extracted_fields.get("business_name"):
                extracted_business_names.append(extracted_fields["business_name"])
    
    # Update run with extracted business name (use first found, or most common)
    if extracted_business_names and not run.get("business_name"):
        # Use the first extracted business name, or most common if multiple
        from collections import Counter
        if len(extracted_business_names) > 1:
            most_common = Counter(extracted_business_names).most_common(1)[0][0]
            run["business_name"] = most_common
        else:
            run["business_name"] = extracted_business_names[0]
        logger.info(f"Extracted business name: {run['business_name']}")
    
    run["extracted"] = True
    save_runs(runs)
    
    logger.info(f"Extracted {len(extracted_docs)} documents for run {run_id}")
    return extracted_docs


# Formatting constants per specification
FILL_HEADER = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
FILL_YELLOW = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
FILL_YELLOW_TOTAL = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
FONT_HEADER = Font(bold=True, size=12, color="FFFFFF")
FONT_TITLE = Font(bold=True, size=14, color="FFFFFF")
FONT_SECTION = Font(bold=True, size=11)
FONT_TOTAL = Font(bold=True, size=12)
FONT_NORMAL = Font(size=11)
ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_RIGHT = Alignment(horizontal="right")
ALIGN_LEFT = Alignment(horizontal="left")

# Market benchmark constants (NSW/Victoria - adjust as needed)
MARKET_BENCHMARKS = {
    "electricity_sme": {
        "peak_rate_c_per_kwh": (25, 32),  # c/kWh range
        "off_peak_rate_c_per_kwh": (18, 24),
        "shoulder_rate_c_per_kwh": (20, 26),
        "daily_supply_charge": (1.20, 1.80),  # $/day
        "metering_annual": (700, 900),  # $/year
        "demand_charge_per_kva_month": (8, 12)  # $/kVA/month
    },
    "electricity_ci": {
        "peak_rate_c_per_kwh": (20, 28),
        "off_peak_rate_c_per_kwh": (15, 22),
        "shoulder_rate_c_per_kwh": (18, 24),
        "daily_supply_charge": (1.50, 5.00),
        "metering_annual": (700, 900),
        "demand_charge_per_kva_month": (10, 15)
    },
    "gas_sme": {
        "rate_per_gj": (16.0, 18.5),  # $/GJ
        "daily_supply_charge": (0.90, 1.20),
        "distribution_per_gj": (2.0, 5.0)
    },
    "gas_ci": {
        "rate_per_gj": (14.0, 16.75),
        "daily_supply_charge": (1.00, 1.50),
        "distribution_per_gj": (2.0, 5.0)
    },
    "waste": {
        "general_compactor_per_changeover": (180, 250),
        "general_bins_per_collection": (15, 30),
        "cardboard_recycling_per_collection": (0, 15),
        "organic_waste_per_collection": (15, 25),
        "fuel_surcharge_percent": (5, 15)
    },
    "water_sydney": {
        "water_usage_per_kl": (2.50, 2.90),
        "wastewater_usage_per_kl": (1.30, 1.50),
        "trade_waste_per_kl": (2.00, 2.50)
    }
}


def safe_float(value, default=0.0):
    """Safely convert value to float"""
    if value is None or value == "":
        return default
    try:
        return float(str(value).replace(",", "").replace("$", "").replace(" ", ""))
    except (ValueError, TypeError):
        return default


def safe_str(value, default=""):
    """Safely convert value to string"""
    if value is None:
        return default
    return str(value).strip() if str(value).strip() else default


def format_currency(value):
    """Format value as currency"""
    return f"${safe_float(value):,.2f}"


def format_number(value):
    """Format value as number with commas"""
    return f"{safe_float(value):,.2f}"


def format_date(date_str):
    """Format date string to DD MMM YY"""
    if not date_str:
        return ""
    try:
        # Try various date formats
        from datetime import datetime
        formats = ["%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d/%m/%y"]
        for fmt in formats:
            try:
                dt = datetime.strptime(str(date_str).strip(), fmt)
                return dt.strftime("%d %b %y")
            except:
                continue
        return str(date_str)
    except:
        return str(date_str)


def get_billing_period_days(extracted):
    """Get billing period in days"""
    days = extracted.get("billing_days")
    if days:
        try:
            return int(float(str(days)))
        except:
            pass
    
    # Try to calculate from dates
    start = extracted.get("billing_period_start")
    end = extracted.get("billing_period_end")
    if start and end:
        try:
            from datetime import datetime
            formats = ["%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"]
            for fmt in formats:
                try:
                    start_dt = datetime.strptime(str(start).strip(), fmt)
                    end_dt = datetime.strptime(str(end).strip(), fmt)
                    delta = (end_dt - start_dt).days
                    if delta > 0:
                        return delta
                except:
                    continue
        except:
            pass
    
    return None


def group_documents_by_utility(run):
    """Group documents by utility type and account identifier"""
    grouped = {
        "electricity": {},  # key: NMI
        "gas": {},  # key: MIRN
        "water": {},  # key: account_number
        "waste": {},  # key: account_number
        "oil": {},  # key: account_number
        "cleaning": {},  # key: account_number
    }
    
    sites = set()
    
    for doc in run.get("documents", []):
        extracted = doc.get("extracted_fields", {})
        if not extracted or extracted.get("error"):
            continue
        
        utility_type = extracted.get("utility_type", "").lower()
        site_address = extracted.get("site_address", "")
        if site_address:
            sites.add(site_address)
        
        if "electricity" in utility_type:
            nmi = extracted.get("nmi") or extracted.get("account_number", "")
            if nmi:
                if nmi not in grouped["electricity"]:
                    grouped["electricity"][nmi] = []
                grouped["electricity"][nmi].append(extracted)
        elif "gas" in utility_type:
            mrin = extracted.get("mrin") or extracted.get("account_number", "")
            if mrin:
                if mrin not in grouped["gas"]:
                    grouped["gas"][mrin] = []
                grouped["gas"][mrin].append(extracted)
        elif "water" in utility_type:
            account = extracted.get("account_number", "") or "Unknown"
            if account not in grouped["water"]:
                grouped["water"][account] = []
            grouped["water"][account].append(extracted)
        elif "waste" in utility_type:
            account = extracted.get("account_number", "") or "Unknown"
            if account not in grouped["waste"]:
                grouped["waste"][account] = []
            grouped["waste"][account].append(extracted)
        elif "oil" in utility_type:
            account = extracted.get("account_number", "") or "Unknown"
            if account not in grouped["oil"]:
                grouped["oil"][account] = []
            grouped["oil"][account].append(extracted)
        elif "cleaning" in utility_type:
            account = extracted.get("account_number", "") or "Unknown"
            if account not in grouped["cleaning"]:
                grouped["cleaning"][account] = []
            grouped["cleaning"][account].append(extracted)
    
    return grouped, list(sites)


def generate_excel_workbook(run_id: str) -> str:
    """Generate Base1 Excel workbook with 7 mandatory sheets per specification"""
    runs = get_runs()
    if run_id not in runs:
        raise ValueError(f"Run {run_id} not found")
    
    run = runs[run_id]
    run_dir = BASE1_STORAGE / run_id / "outputs"
    run_dir.mkdir(parents=True, exist_ok=True)
    
    wb = Workbook()
    
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    
    # Get business name from first invoice or run
    business_name = run.get("business_name", "")
    if not business_name:
        for doc in run.get("documents", []):
            extracted = doc.get("extracted_fields", {})
            if extracted and extracted.get("business_name"):
                business_name = extracted["business_name"]
                break
    if not business_name:
        business_name = "Business Name"
    
    # Group documents by utility type
    grouped_data, sites = group_documents_by_utility(run)
    
    # Calculate summary totals
    summary_totals = calculate_summary_totals(grouped_data)
    
    # Create all 7 sheets
    create_overview_sheet(wb, business_name, sites, summary_totals, grouped_data, datetime.now())
    create_electricity_sheet(wb, grouped_data["electricity"])
    create_gas_sheet(wb, grouped_data["gas"])
    create_waste_sheet(wb, grouped_data["waste"])
    create_water_sheet(wb, grouped_data["water"])
    create_cost_summary_sheet(wb, business_name, grouped_data, summary_totals)
    create_meter_details_sheet(wb, grouped_data)
    
    # Create Base 1 Analysis sheet (Phase 2) - Savings Opportunities & Benchmarking
    analysis_results = analyze_savings_opportunities(grouped_data, summary_totals)
    create_analysis_sheet(wb, business_name, analysis_results)
    
    # Save workbook
    date_str = datetime.now().strftime("%Y_%m_%d")
    # Clean business name for filename
    clean_name = business_name.replace(" ", "_").replace("PTY", "").replace("LTD", "").replace("Limited", "").replace("Pty", "").replace("Ltd", "").strip("_")
    excel_path = run_dir / f"{clean_name}_Utility_Extraction_{date_str}.xlsx"
    wb.save(excel_path)
    
    logger.info(f"Generated Excel workbook: {excel_path}")
    return str(excel_path)


def calculate_summary_totals(grouped_data):
    """Calculate summary totals for all utilities"""
    totals = {
        "electricity_monthly": 0,
        "gas_monthly": 0,
        "waste_monthly": 0,
        "water_monthly": 0,
        "total_monthly": 0,
                        "total_annual": 0
                    }
    
    # Electricity
    for nmi, invoices in grouped_data["electricity"].items():
        for invoice in invoices:
            total = safe_float(invoice.get("total_inc_gst", 0))
            days = get_billing_period_days(invoice) or 30
            monthly = total * (30 / days) if days > 0 else total
            totals["electricity_monthly"] += monthly
    
    # Gas
    for mrin, invoices in grouped_data["gas"].items():
        for invoice in invoices:
            total = safe_float(invoice.get("total_inc_gst", 0))
            days = get_billing_period_days(invoice) or 30
            monthly = total * (30 / days) if days > 0 else total
            totals["gas_monthly"] += monthly
    
    # Waste
    for account, invoices in grouped_data["waste"].items():
        for invoice in invoices:
            total = safe_float(invoice.get("total_inc_gst", 0))
            days = get_billing_period_days(invoice) or 30
            monthly = total * (30 / days) if days > 0 else total
            totals["waste_monthly"] += monthly
    
    # Water (usually quarterly, divide by 3)
    for account, invoices in grouped_data["water"].items():
        for invoice in invoices:
            total = safe_float(invoice.get("total_inc_gst", 0))
            days = get_billing_period_days(invoice) or 90
            monthly = total * (30 / days) if days > 0 else total
            totals["water_monthly"] += monthly
    
    totals["total_monthly"] = (
        totals["electricity_monthly"] + 
        totals["gas_monthly"] + 
        totals["waste_monthly"] + 
        totals["water_monthly"]
    )
    totals["total_annual"] = totals["total_monthly"] * 12
    
    return totals


def create_overview_sheet(wb, business_name, sites, totals, grouped_data, created_date):
    """Create Overview sheet (Sheet 1)"""
    sheet = wb.create_sheet("Overview", 0)
    
    # Header Block
    row = 1
    sheet.merge_cells(f'A{row}:B{row}')
    cell = sheet.cell(row, 1, f"BASE 1 REVIEW - {business_name.upper()}")
    cell.font = FONT_TITLE
    cell.fill = FILL_HEADER
    cell.alignment = ALIGN_CENTER
    
    row += 1
    sheet.merge_cells(f'A{row}:B{row}')
    sheet.cell(row, 1, f"PROJECT NAME: Base 1 Review - {business_name}")
    
    row += 1
    sheet.merge_cells(f'A{row}:B{row}')
    sheet.cell(row, 1, f"CREATED: {created_date.strftime('%d %b %Y')}")
    
    row += 1
    sheet.merge_cells(f'A{row}:B{row}')
    sheet.cell(row, 1, "PURPOSE: Comprehensive extraction and preliminary analysis of utility expenses")
    
    row += 2
    
    # Client Information
    sheet.cell(row, 1, "Client Information").font = FONT_SECTION
    row += 1
    sheet.cell(row, 1, f"Client Name: {business_name}")
    row += 1
    
    for i, site in enumerate(sites[:10], 1):  # Limit to 10 sites
        sheet.cell(row, 1, f"Site {i}: {site}")
        row += 1
    
    row += 1
    
    # Current Costs Summary
    sheet.cell(row, 1, "Current Costs Summary").font = FONT_SECTION
    row += 1
    sheet.cell(row, 1, f"Electricity (all sites): {format_currency(totals['electricity_monthly'])}/month")
    row += 1
    sheet.cell(row, 1, f"Gas: {format_currency(totals['gas_monthly'])}/month")
    row += 1
    sheet.cell(row, 1, f"Waste: {format_currency(totals['waste_monthly'])}/month")
    row += 1
    sheet.cell(row, 1, f"Water & Wastewater: {format_currency(totals['water_monthly'])}/month")
    row += 1
    sheet.cell(row, 1, f"ESTIMATED TOTAL MONTHLY SPEND: {format_currency(totals['total_monthly'])}/month").font = FONT_TOTAL
    row += 1
    sheet.cell(row, 1, f"ESTIMATED ANNUAL SPEND: {format_currency(totals['total_annual'])}/year").font = FONT_TOTAL
    
    row += 2
    
    # Quick Facts
    sheet.cell(row, 1, "Quick Facts").font = FONT_SECTION
    row += 1
    facts = []
    if totals['electricity_monthly'] > 0:
        facts.append(f"✅ Electricity accounts identified: {len(grouped_data.get('electricity', {}))}")
    if totals['gas_monthly'] > 0:
        facts.append(f"✅ Gas accounts identified: {len(grouped_data.get('gas', {}))}")
    if totals['waste_monthly'] > 0:
        facts.append(f"🗑️ Waste services identified: {len(grouped_data.get('waste', {}))}")
    if totals['water_monthly'] > 0:
        facts.append(f"💧 Water accounts identified: {len(grouped_data.get('water', {}))}")
    if totals['total_annual'] > 100000:
        facts.append(f"💡 Significant annual spend identified - potential for substantial savings")
    if len(sites) > 1:
        facts.append(f"📊 Multiple sites identified: {len(sites)} locations")
    
    for fact in facts[:6]:  # Limit to 6 facts
        sheet.cell(row, 1, fact)
        row += 1
    
    # Add savings highlights if analysis is available
    # (This will be populated when analysis is run)
    row += 1
    sheet.cell(row, 1, "💡 See 'Base 1 Analysis' sheet for detailed savings opportunities and benchmarking")
    
    row += 2
    
    # How to Use This Workbook
    sheet.cell(row, 1, "How to Use This Workbook").font = FONT_SECTION
    row += 1
    sheet.cell(row, 1, "1. Overview - This summary page")
    row += 1
    sheet.cell(row, 1, "2. Electricity Data - All electricity accounts detailed")
    row += 1
    sheet.cell(row, 1, "3. Gas Data - Gas/LPG account details and breakdown")
    row += 1
    sheet.cell(row, 1, "4. Waste Data - Comprehensive waste management costs")
    row += 1
    sheet.cell(row, 1, "5. Water Data - Water, wastewater & trade waste charges")
    row += 1
    sheet.cell(row, 1, "6. Cost Summary - Consolidated cost analysis")
    row += 1
    sheet.cell(row, 1, "7. Meter Details - All meter numbers and readings")
    
    # Set column widths
    sheet.column_dimensions['A'].width = 50
    sheet.column_dimensions['B'].width = 60


def create_electricity_sheet(wb, electricity_data):
    """Create Electricity Data sheet (Sheet 2)"""
    sheet = wb.create_sheet("Electricity Data", 1)
    
    # Headers
    headers = [
        "Site", "Account Details", "NMI/Meter", "Billing Period", "Invoice Date",
        "Invoice #", "Consumption (kWh)", "Peak (kWh)", "Shoulder (kWh)", "Off-Peak (kWh)",
        "Demand (kVA/kW)", "Total Amount ($)", "Supplier", "Tariff", "Notes"
    ]
    sheet.append(headers)
    
    # Format headers
    for col in range(1, len(headers) + 1):
        cell = sheet.cell(1, col)
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.alignment = ALIGN_CENTER
    
    # Column widths
    widths = [18, 25, 18, 25, 15, 18, 15, 12, 12, 12, 15, 15, 18, 15, 50]
    for i, width in enumerate(widths, 1):
        sheet.column_dimensions[chr(64 + i)].width = width
    
    # Data rows
    row = 2
    total_kwh = 0
    total_cost = 0
    
    for nmi, invoices in electricity_data.items():
        for invoice in invoices:
            site = safe_str(invoice.get("site_address", ""))
            account = safe_str(invoice.get("account_number", ""))
            billing_start = safe_str(invoice.get("billing_period_start", ""))
            billing_end = safe_str(invoice.get("billing_period_end", ""))
            days = get_billing_period_days(invoice)
            billing_period = f"{billing_start} - {billing_end} ({days} days)" if billing_start and billing_end and days else safe_str(invoice.get("billing_period", ""))
            
            invoice_date = format_date(invoice.get("invoice_date", ""))
            invoice_num = safe_str(invoice.get("invoice_number", ""))
            
            peak_kwh = safe_float(invoice.get("peak_usage_kwh", 0))
            shoulder_kwh = safe_float(invoice.get("shoulder_usage_kwh", 0))
            off_peak_kwh = safe_float(invoice.get("off_peak_usage_kwh", 0))
            total_usage = peak_kwh + shoulder_kwh + off_peak_kwh
            if total_usage == 0:
                total_usage = safe_float(invoice.get("total_usage_kwh", 0))
            
            demand = safe_str(invoice.get("demand_kw", "")) or safe_str(invoice.get("demand_kva", ""))
            total_amount = safe_float(invoice.get("total_inc_gst", 0))
            supplier = safe_str(invoice.get("supplier", ""))
            tariff = safe_str(invoice.get("tariff_type", ""))
            
            # Notes
            notes_parts = []
            if safe_float(invoice.get("meter_charges", 0)) > 0:
                notes_parts.append(f"Metering: {format_currency(invoice.get('meter_charges', 0))}")
            if safe_float(invoice.get("network_charges_ex_gst", 0)) > 0:
                notes_parts.append(f"Network: {format_currency(invoice.get('network_charges_ex_gst', 0))}")
            notes = "; ".join(notes_parts) if notes_parts else ""
            
            sheet.append([
                site, account, nmi, billing_period, invoice_date, invoice_num,
                format_number(total_usage), format_number(peak_kwh), format_number(shoulder_kwh),
                format_number(off_peak_kwh), demand, format_currency(total_amount),
                supplier, tariff, notes
            ])
            
            total_kwh += total_usage
            total_cost += total_amount
            row += 1
    
    # Totals row
    if row > 2:
        sheet.append([
            "TOTALS:", "", "", "", "", "",
            format_number(total_kwh), "", "", "",
            "", format_currency(total_cost), "", "",
            f"Average rate: {format_currency(total_cost / total_kwh if total_kwh > 0 else 0)}/kWh"
        ])
        # Format totals row
        for col in [1, 7, 12, 15]:
            cell = sheet.cell(row, col)
            cell.font = FONT_TOTAL
            if col == 12:  # Amount column
                cell.fill = FILL_YELLOW


def create_gas_sheet(wb, gas_data):
    """Create Gas Data sheet (Sheet 3)"""
    sheet = wb.create_sheet("Gas Data", 2)
    
    # Headers
    headers = [
        "Site", "Account Number", "MIRN", "Meter Number", "Billing Period",
        "Invoice Date", "Consumption (GJ)", "Consumption (m³)", "PCF", "HV",
        "Total Amount ($)", "Energy Charges", "Transmission", "Distribution",
        "Other Charges", "Supplier"
    ]
    sheet.append(headers)
    
    # Format headers
    for col in range(1, len(headers) + 1):
        cell = sheet.cell(1, col)
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.alignment = ALIGN_CENTER
    
    # Column widths
    widths = [25, 18, 18, 18, 28, 15, 15, 15, 8, 8, 15, 15, 15, 15, 15, 20]
    for i, width in enumerate(widths, 1):
        sheet.column_dimensions[chr(64 + i)].width = width
    
    # Data rows
    row = 2
    total_gj = 0
    total_m3 = 0
    total_amount = 0
    
    for mrin, invoices in gas_data.items():
        for invoice in invoices:
            site = safe_str(invoice.get("site_address", ""))
            account = safe_str(invoice.get("account_number", ""))
            meter_num = safe_str(invoice.get("meter_number", ""))
            
            billing_start = safe_str(invoice.get("billing_period_start", ""))
            billing_end = safe_str(invoice.get("billing_period_end", ""))
            days = get_billing_period_days(invoice)
            billing_period = f"{billing_start} - {billing_end} ({days} days)" if billing_start and billing_end and days else safe_str(invoice.get("billing_period", ""))
            
            invoice_date = format_date(invoice.get("invoice_date", ""))
            
            consumption_gj = safe_float(invoice.get("total_usage_gj", 0))
            consumption_m3 = safe_float(invoice.get("volume_m3", 0))
            pcf = safe_str(invoice.get("pcf", ""))
            hv = safe_str(invoice.get("hv", ""))
            
            total_amt = safe_float(invoice.get("total_inc_gst", 0))
            energy_charges = safe_float(invoice.get("usage_charges_ex_gst", 0))
            transmission = safe_float(invoice.get("transmission_charges", 0))
            distribution = safe_float(invoice.get("distribution_charges", 0))
            other_charges = total_amt - energy_charges - transmission - distribution
            supplier = safe_str(invoice.get("supplier", ""))
            
            sheet.append([
                site, account, mrin, meter_num, billing_period, invoice_date,
                format_number(consumption_gj), format_number(consumption_m3), pcf, hv,
                format_currency(total_amt), format_currency(energy_charges),
                format_currency(transmission), format_currency(distribution),
                format_currency(other_charges), supplier
            ])
            
            total_gj += consumption_gj
            total_m3 += consumption_m3
            total_amount += total_amt
            row += 1
    
    # Totals row
    if row > 2:
        sheet.append([
            "TOTALS:", "", "", "", "", "",
            format_number(total_gj), format_number(total_m3), "", "",
            format_currency(total_amount), format_currency(total_amount * 0.7),  # Estimate breakdowns
            format_currency(total_amount * 0.15), format_currency(total_amount * 0.10),
            format_currency(total_amount * 0.05), ""
        ])
        # Format totals row
        for col in [1, 7, 8, 11, 12, 13, 14, 15]:
            cell = sheet.cell(row, col)
            cell.font = FONT_TOTAL


def create_waste_sheet(wb, waste_data):
    """Create Waste Data sheet (Sheet 4)"""
    sheet = wb.create_sheet("Waste Data", 3)
    
    # Headers
    headers = [
        "Site", "Service Type", "Bin Type/Size", "Collections per Month",
        "Rate per Collection", "Monthly Cost ($)", "Provider", "Account Number", "Details/Notes"
    ]
    sheet.append(headers)
    
    # Format headers
    for col in range(1, len(headers) + 1):
        cell = sheet.cell(1, col)
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.alignment = ALIGN_CENTER
    
    # Column widths
    widths = [25, 22, 20, 22, 20, 18, 15, 18, 60]
    for i, width in enumerate(widths, 1):
        sheet.column_dimensions[chr(64 + i)].width = width
    
    # Data rows
    row = 2
    subtotal = 0
    
    for account, invoices in waste_data.items():
        for invoice in invoices:
            site = safe_str(invoice.get("site_address", ""))
            service_type = safe_str(invoice.get("service_type", "General Waste"))
            bin_type = safe_str(invoice.get("bin_type", "")) or safe_str(invoice.get("bin_size", ""))
            
            # Try to extract collections per month
            collections = safe_str(invoice.get("collections_per_month", ""))
            if not collections:
                # Estimate from billing period
                days = get_billing_period_days(invoice) or 30
                total_collections = safe_float(invoice.get("total_collections", 1))
                collections = f"{total_collections * (30 / days):.1f}" if days > 0 else "1"
            
            rate_per_collection = safe_str(invoice.get("rate_per_collection", ""))
            monthly_cost = safe_float(invoice.get("total_inc_gst", 0))
            days = get_billing_period_days(invoice) or 30
            if days != 30:
                monthly_cost = monthly_cost * (30 / days) if days > 0 else monthly_cost
            
            provider = safe_str(invoice.get("supplier", ""))
            details = safe_str(invoice.get("notes", "")) or safe_str(invoice.get("details", ""))
            
            sheet.append([
                site, service_type, bin_type, collections, rate_per_collection,
                format_currency(monthly_cost), provider, account, details
            ])
            
            subtotal += monthly_cost
            row += 1
    
    # Totals section
    if row > 2:
        row += 1
        sheet.cell(row, 5, "MONTHLY SUBTOTAL:").font = FONT_TOTAL
        sheet.cell(row, 6, format_currency(subtotal)).font = FONT_TOTAL
        sheet.cell(row, 6).fill = FILL_YELLOW
        
        row += 1
        gst = subtotal * 0.10
        sheet.cell(row, 5, "GST (10%):").font = FONT_TOTAL
        sheet.cell(row, 6, format_currency(gst)).font = FONT_TOTAL
        
        row += 1
        total = subtotal + gst
        sheet.cell(row, 5, "TOTAL MONTHLY WASTE:").font = FONT_TOTAL
        sheet.cell(row, 6, format_currency(total)).font = FONT_TOTAL
        sheet.cell(row, 6).fill = FILL_YELLOW


def create_water_sheet(wb, water_data):
    """Create Water Data sheet (Sheet 5)"""
    sheet = wb.create_sheet("Water Data", 4)
    
    # Headers
    headers = [
        "Site", "Account Number", "Service Type", "Billing Period",
        "Invoice Date", "Quantity/Usage", "Rate", "Amount ($)", "Details"
    ]
    sheet.append(headers)
    
    # Format headers
    for col in range(1, len(headers) + 1):
        cell = sheet.cell(1, col)
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.alignment = ALIGN_CENTER
    
    # Column widths
    widths = [25, 18, 35, 25, 15, 18, 15, 15, 50]
    for i, width in enumerate(widths, 1):
        sheet.column_dimensions[chr(64 + i)].width = width
    
    # Data rows
    row = 2
    quarterly_total = 0
    
    for account, invoices in water_data.items():
        for invoice in invoices:
            site = safe_str(invoice.get("site_address", ""))
            service_type = safe_str(invoice.get("service_type", "Water Service (Fixed)"))
            
            billing_start = safe_str(invoice.get("billing_period_start", ""))
            billing_end = safe_str(invoice.get("billing_period_end", ""))
            days = get_billing_period_days(invoice)
            billing_period = f"{billing_start} - {billing_end} ({days} days)" if billing_start and billing_end and days else safe_str(invoice.get("billing_period", ""))
            
            invoice_date = format_date(invoice.get("invoice_date", ""))
            quantity = safe_str(invoice.get("quantity", "")) or safe_str(invoice.get("usage", ""))
            rate = safe_str(invoice.get("rate", ""))
            amount = safe_float(invoice.get("total_inc_gst", 0))
            details = safe_str(invoice.get("details", "")) or safe_str(invoice.get("notes", ""))
            
            sheet.append([
                site, account, service_type, billing_period, invoice_date,
                quantity, rate, format_currency(amount), details
            ])
            
            quarterly_total += amount
            row += 1
    
    # Totals section
    if row > 2:
        row += 1
        sheet.cell(row, 7, "TOTAL WATER & WASTEWATER:").font = FONT_TOTAL
        sheet.cell(row, 8, format_currency(quarterly_total)).font = FONT_TOTAL
        sheet.cell(row, 8).fill = FILL_YELLOW
        sheet.cell(row, 9, "Quarterly bill")
        
        row += 1
        monthly_estimate = quarterly_total / 3
        sheet.cell(row, 7, "ESTIMATED MONTHLY COST:").font = FONT_TOTAL
        sheet.cell(row, 8, format_currency(monthly_estimate)).font = FONT_TOTAL
        sheet.cell(row, 8).fill = FILL_YELLOW
        sheet.cell(row, 9, "Quarterly ÷ 3")


def create_cost_summary_sheet(wb, business_name, grouped_data, totals):
    """Create Cost Summary sheet (Sheet 6)"""
    sheet = wb.create_sheet("Cost Summary", 5)
    
    # Title
    sheet.merge_cells('A1:F1')
    cell = sheet.cell(1, 1, f"{business_name.upper()} - COST SUMMARY")
    cell.font = FONT_TITLE
    cell.fill = FILL_HEADER
    cell.alignment = ALIGN_CENTER
    
    row = 3
    
    # Headers
    headers = ["Utility Type", "Site", "Monthly Cost ($)", "Quarterly Cost ($)", "Annual Estimate ($)", "Notes"]
    sheet.append(headers)
    
    # Format headers
    for col in range(1, len(headers) + 1):
        cell = sheet.cell(row, col)
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.alignment = ALIGN_CENTER
    
    # Column widths
    widths = [30, 30, 20, 20, 20, 40]
    for i, width in enumerate(widths, 1):
        sheet.column_dimensions[chr(64 + i)].width = width
    
    row = 4
    
    # Electricity rows
    for nmi, invoices in grouped_data["electricity"].items():
        monthly = 0
        for invoice in invoices:
            total = safe_float(invoice.get("total_inc_gst", 0))
            days = get_billing_period_days(invoice) or 30
            monthly += total * (30 / days) if days > 0 else total
        
        site = safe_str(invoices[0].get("site_address", "")) if invoices else ""
        sheet.append([
            "Electricity", site, format_currency(monthly),
            format_currency(monthly * 3), format_currency(monthly * 12),
            f"NMI: {nmi}"
        ])
        row += 1
    
    # Gas rows
    for mrin, invoices in grouped_data["gas"].items():
        monthly = 0
        for invoice in invoices:
            total = safe_float(invoice.get("total_inc_gst", 0))
            days = get_billing_period_days(invoice) or 30
            monthly += total * (30 / days) if days > 0 else total
        
        site = safe_str(invoices[0].get("site_address", "")) if invoices else ""
        sheet.append([
            "Gas", site, format_currency(monthly),
            format_currency(monthly * 3), format_currency(monthly * 12),
            f"MIRN: {mrin}"
        ])
        row += 1
    
    # Waste rows
    for account, invoices in grouped_data["waste"].items():
        monthly = 0
        for invoice in invoices:
            total = safe_float(invoice.get("total_inc_gst", 0))
            days = get_billing_period_days(invoice) or 30
            monthly += total * (30 / days) if days > 0 else total
        
        site = safe_str(invoices[0].get("site_address", "")) if invoices else ""
        sheet.append([
            "Waste", site, format_currency(monthly),
            format_currency(monthly * 3), format_currency(monthly * 12),
            f"Account: {account}"
        ])
        row += 1
    
    # Water rows
    for account, invoices in grouped_data["water"].items():
        monthly = 0
        for invoice in invoices:
            total = safe_float(invoice.get("total_inc_gst", 0))
            days = get_billing_period_days(invoice) or 90  # Usually quarterly
            monthly += total * (30 / days) if days > 0 else total
        
        site = safe_str(invoices[0].get("site_address", "")) if invoices else ""
        sheet.append([
            "Water & Wastewater", site, format_currency(monthly),
            format_currency(monthly * 3), format_currency(monthly * 12),
            f"Account: {account}"
        ])
        row += 1
    
    # Totals row
    row += 1
    sheet.append([
        "TOTALS:", "", format_currency(totals["total_monthly"]),
        format_currency(totals["total_monthly"] * 3), format_currency(totals["total_annual"]), ""
    ])
    # Format totals row
    for col in [1, 3, 4, 5]:
        cell = sheet.cell(row, col)
        cell.font = FONT_TOTAL
        if col in [3, 4, 5]:
            cell.fill = FILL_YELLOW_TOTAL
    
    # Breakdown by utility type
    row += 2
    sheet.cell(row, 1, "BREAKDOWN BY UTILITY TYPE").font = FONT_SECTION
    row += 1
    
    if totals["total_annual"] > 0:
        utilities = [
            ("Electricity (all sites)", totals["electricity_monthly"] * 12),
            ("Gas", totals["gas_monthly"] * 12),
            ("Waste", totals["waste_monthly"] * 12),
            ("Water & Wastewater", totals["water_monthly"] * 12)
        ]
        
        for util_name, annual_cost in utilities:
            if annual_cost > 0:
                percentage = (annual_cost / totals["total_annual"]) * 100
                sheet.append([
                    util_name, "", "", "", format_currency(annual_cost),
                    f"{percentage:.1f}%"
                ])
                row += 1


def analyze_savings_opportunities(grouped_data, totals):
    """Analyze all utility data and identify savings opportunities"""
    opportunities = []
    critical_issues = []
    warnings = []
    
    # Analyze electricity
    for nmi, invoices in grouped_data["electricity"].items():
        for invoice in invoices:
            opps = analyze_electricity_invoice(invoice)
            opportunities.extend(opps)
            
            # Check for critical issues
            for opp in opps:
                if opp.get("severity") == "critical":
                    critical_issues.append(opp)
                elif opp.get("severity") == "high":
                    warnings.append(opp)
    
    # Analyze gas
    for mrin, invoices in grouped_data["gas"].items():
        for invoice in invoices:
            opps = analyze_gas_invoice(invoice)
            opportunities.extend(opps)
            for opp in opps:
                if opp.get("severity") == "critical":
                    critical_issues.append(opp)
                elif opp.get("severity") == "high":
                    warnings.append(opp)
    
    # Analyze waste
    waste_opps = analyze_waste_opportunities(grouped_data["waste"])
    opportunities.extend(waste_opps)
    
    # Calculate potential savings
    total_potential_savings = calculate_potential_savings(opportunities, grouped_data, totals)
    
    return {
        "opportunities": opportunities,
        "critical_issues": critical_issues,
        "warnings": warnings,
        "total_potential_savings": total_potential_savings,
        "immediate_actions": [o for o in opportunities if o.get("timeframe") == "0-30 days"],
        "short_term_actions": [o for o in opportunities if o.get("timeframe") == "1-3 months"]
    }


def analyze_electricity_invoice(invoice):
    """Analyze electricity invoice for savings opportunities"""
    opportunities = []
    utility_type = invoice.get("utility_type", "").lower()
    is_ci = "c&i" in utility_type or "ci" in utility_type or "commercial" in utility_type or "industrial" in utility_type
    benchmarks = MARKET_BENCHMARKS["electricity_ci" if is_ci else "electricity_sme"]
    
    # Check peak rate
    peak_rate = invoice.get("peak_rate_c_per_kwh")
    if peak_rate:
        peak_val = safe_float(peak_rate)
        peak_range = benchmarks["peak_rate_c_per_kwh"]
        if peak_val > peak_range[1]:
            annual_savings = calculate_electricity_rate_savings(invoice, "peak", peak_val, peak_range[1])
            opportunities.append({
                "category": "Electricity",
                "type": "High Peak Rate",
                "severity": "high",
                "flag": "⚠️ WARNING",
                "issue": f"Peak rate {peak_val:.2f} c/kWh exceeds market benchmark ({peak_range[0]}-{peak_range[1]} c/kWh)",
                "current_rate": f"{peak_val:.2f} c/kWh",
                "market_benchmark": f"{peak_range[0]}-{peak_range[1]} c/kWh",
                "potential_savings_annual": annual_savings,
                "recommendation": "Consider competitive tender or negotiate with current retailer",
                "timeframe": "1-3 months"
            })
        elif peak_val > peak_range[1] * 0.9:  # Within 10% of upper limit
            opportunities.append({
                "category": "Electricity",
                "type": "Elevated Peak Rate",
                "severity": "medium",
                "flag": "💡 OPPORTUNITY",
                "issue": f"Peak rate {peak_val:.2f} c/kWh is at higher end of market range",
                "current_rate": f"{peak_val:.2f} c/kWh",
                "market_benchmark": f"{peak_range[0]}-{peak_range[1]} c/kWh",
                "potential_savings_annual": calculate_electricity_rate_savings(invoice, "peak", peak_val, peak_range[1] * 0.9),
                "recommendation": "Worth reviewing - may be able to negotiate better rate",
                "timeframe": "1-3 months"
            })
    
    # Check off-peak rate
    off_peak_rate = invoice.get("off_peak_rate_c_per_kwh")
    if off_peak_rate:
        off_peak_val = safe_float(off_peak_rate)
        off_peak_range = benchmarks["off_peak_rate_c_per_kwh"]
        if off_peak_val > off_peak_range[1]:
            annual_savings = calculate_electricity_rate_savings(invoice, "off_peak", off_peak_val, off_peak_range[1])
            opportunities.append({
                "category": "Electricity",
                "type": "High Off-Peak Rate",
                "severity": "medium",
                "flag": "💡 OPPORTUNITY",
                "issue": f"Off-peak rate {off_peak_val:.2f} c/kWh exceeds market benchmark",
                "current_rate": f"{off_peak_val:.2f} c/kWh",
                "market_benchmark": f"{off_peak_range[0]}-{off_peak_range[1]} c/kWh",
                "potential_savings_annual": annual_savings,
                "recommendation": "Review off-peak rates with retailer",
                "timeframe": "1-3 months"
            })
    
    # Check metering charges
    metering_charges = invoice.get("meter_charges")
    if metering_charges:
        metering_val = safe_float(metering_charges)
        days = get_billing_period_days(invoice) or 30
        monthly_metering = metering_val * (30 / days) if days > 0 else metering_val
        annual_metering = monthly_metering * 12
        metering_range = benchmarks["metering_annual"]
        
        if annual_metering > metering_range[1]:
            savings = annual_metering - metering_range[1]
            opportunities.append({
                "category": "Electricity",
                "type": "High Metering Charges",
                "severity": "high",
                "flag": "⚠️ WARNING",
                "issue": f"Metering charges ${annual_metering:,.2f}/year exceed market benchmark (${metering_range[0]}-${metering_range[1]}/year)",
                "current_cost": format_currency(annual_metering),
                "market_benchmark": f"${metering_range[0]}-{metering_range[1]}/year",
                "potential_savings_annual": savings,
                "recommendation": "Verify not double-charged, consider alternative metering provider",
                "timeframe": "0-30 days"
            })
    
    # Check demand charges
    demand_charges = invoice.get("demand_charges")
    demand_kw = invoice.get("demand_kw") or invoice.get("demand_kva")
    if demand_charges and demand_kw:
        demand_charges_val = safe_float(demand_charges)
        demand_kw_val = safe_float(demand_kw)
        days = get_billing_period_days(invoice) or 30
        monthly_demand = demand_charges_val * (30 / days) if days > 0 else demand_charges_val
        
        if demand_kw_val > 0:
            rate_per_kva_month = monthly_demand / demand_kw_val
            demand_range = benchmarks["demand_charge_per_kva_month"]
            
            if rate_per_kva_month > demand_range[1]:
                annual_demand = monthly_demand * 12
                benchmark_annual = demand_kw_val * demand_range[1] * 12
                savings = annual_demand - benchmark_annual
                opportunities.append({
                    "category": "Electricity",
                    "type": "High Demand Charges",
                    "severity": "high",
                    "flag": "⚠️ WARNING",
                    "issue": f"Demand charges ${rate_per_kva_month:.2f}/kVA/month exceed market benchmark (${demand_range[0]}-${demand_range[1]}/kVA/month)",
                    "current_rate": f"${rate_per_kva_month:.2f}/kVA/month",
                    "market_benchmark": f"${demand_range[0]}-{demand_range[1]}/kVA/month",
                    "potential_savings_annual": savings,
                    "recommendation": "Implement demand management strategies, review tariff structure",
                    "timeframe": "1-3 months"
                })
    
    # Check daily supply charge
    daily_supply = invoice.get("daily_supply_charge")
    if daily_supply:
        supply_val = safe_float(daily_supply)
        supply_range = benchmarks["daily_supply_charge"]
        if supply_val > supply_range[1]:
            annual_supply = supply_val * 365
            benchmark_annual = supply_range[1] * 365
            savings = annual_supply - benchmark_annual
            opportunities.append({
                "category": "Electricity",
                "type": "High Supply Charge",
                "severity": "medium",
                "flag": "💡 OPPORTUNITY",
                "issue": f"Daily supply charge ${supply_val:.2f}/day may be above market rates",
                "current_rate": f"${supply_val:.2f}/day",
                "market_benchmark": f"${supply_range[0]}-{supply_range[1]}/day",
                "potential_savings_annual": savings,
                "recommendation": "Review supply charges with retailer",
                "timeframe": "1-3 months"
            })
    
    return opportunities


def analyze_gas_invoice(invoice):
    """Analyze gas invoice for savings opportunities"""
    opportunities = []
    utility_type = invoice.get("utility_type", "").lower()
    is_ci = "c&i" in utility_type or "ci" in utility_type or "commercial" in utility_type or "industrial" in utility_type
    benchmarks = MARKET_BENCHMARKS["gas_ci" if is_ci else "gas_sme"]
    
    # Check gas rate
    total_usage_gj = safe_float(invoice.get("total_usage_gj", 0))
    if total_usage_gj == 0:
        # Try to convert from MJ
        total_usage_mj = safe_float(invoice.get("total_usage_mj", 0))
        total_usage_gj = total_usage_mj / 1000 if total_usage_mj > 0 else 0
    
    usage_charges = safe_float(invoice.get("usage_charges_ex_gst", 0))
    if usage_charges > 0 and total_usage_gj > 0:
        rate_per_gj = usage_charges / total_usage_gj
        rate_range = benchmarks["rate_per_gj"]
        
        if rate_per_gj > rate_range[1]:
            annual_usage = total_usage_gj * 12  # Estimate annual
            benchmark_cost = annual_usage * rate_range[1]
            current_cost = annual_usage * rate_per_gj
            savings = current_cost - benchmark_cost
            opportunities.append({
                "category": "Gas",
                "type": "High Gas Rate",
                "severity": "high",
                "flag": "⚠️ WARNING",
                "issue": f"Gas rate ${rate_per_gj:.2f}/GJ exceeds market benchmark (${rate_range[0]}-{rate_range[1]}/GJ)",
                "current_rate": f"${rate_per_gj:.2f}/GJ",
                "market_benchmark": f"${rate_range[0]}-{rate_range[1]}/GJ",
                "potential_savings_annual": savings,
                "recommendation": "Consider competitive tender or negotiate with supplier",
                "timeframe": "1-3 months"
            })
    
    # Check daily supply charge
    daily_supply = invoice.get("daily_supply_charge")
    if daily_supply:
        supply_val = safe_float(daily_supply)
        supply_range = benchmarks["daily_supply_charge"]
        if supply_val > supply_range[1]:
            annual_supply = supply_val * 365
            benchmark_annual = supply_range[1] * 365
            savings = annual_supply - benchmark_annual
            opportunities.append({
                "category": "Gas",
                "type": "High Supply Charge",
                "severity": "medium",
                "flag": "💡 OPPORTUNITY",
                "issue": f"Daily supply charge ${supply_val:.2f}/day may be above market rates",
                "current_rate": f"${supply_val:.2f}/day",
                "market_benchmark": f"${supply_range[0]}-{supply_range[1]}/day",
                "potential_savings_annual": savings,
                "recommendation": "Review supply charges with supplier",
                "timeframe": "1-3 months"
            })
    
    return opportunities


def analyze_waste_opportunities(waste_data):
    """Analyze waste data for consolidation and reduction opportunities"""
    opportunities = []
    
    # Check for multiple providers at same location
    sites_providers = {}
    for account, invoices in waste_data.items():
        for invoice in invoices:
            site = safe_str(invoice.get("site_address", ""))
            provider = safe_str(invoice.get("supplier", ""))
            if site and provider:
                if site not in sites_providers:
                    sites_providers[site] = set()
                sites_providers[site].add(provider)
    
    for site, providers in sites_providers.items():
        if len(providers) > 1:
            # Calculate potential savings from consolidation
            site_total = 0
            for account, invoices in waste_data.items():
                for invoice in invoices:
                    if safe_str(invoice.get("site_address", "")) == site:
                        total = safe_float(invoice.get("total_inc_gst", 0))
                        days = get_billing_period_days(invoice) or 30
                        monthly = total * (30 / days) if days > 0 else total
                        site_total += monthly
            
            # Estimate 10-15% savings from consolidation
            potential_savings = site_total * 12 * 0.12  # 12% annual savings estimate
            
            opportunities.append({
                "category": "Waste",
                "type": "Multiple Providers - Consolidate",
                "severity": "high",
                "flag": "🔴 CRITICAL",
                "issue": f"Multiple waste providers ({len(providers)}) at {site} - consolidation opportunity",
                "current_situation": f"{len(providers)} providers at same location",
                "recommendation": "CONSOLIDATE to single provider for better rates and simplified management",
                "potential_savings_annual": potential_savings,
                "timeframe": "0-30 days"
            })
    
    # Check for high collection frequency
    for account, invoices in waste_data.items():
        for invoice in invoices:
            collections = safe_str(invoice.get("collections_per_month", ""))
            if collections:
                try:
                    coll_per_month = float(collections)
                    if coll_per_month > 12:  # More than 3x per week
                        days = get_billing_period_days(invoice) or 30
                        monthly_cost = safe_float(invoice.get("total_inc_gst", 0)) * (30 / days) if days > 0 else safe_float(invoice.get("total_inc_gst", 0))
                        # Estimate 20-30% savings from reducing frequency
                        potential_savings = monthly_cost * 12 * 0.25
                        opportunities.append({
                            "category": "Waste",
                            "type": "High Collection Frequency",
                            "severity": "medium",
                            "flag": "💡 OPPORTUNITY",
                            "issue": f"Collection frequency {coll_per_month:.1f}/month may be excessive",
                            "current_situation": f"{coll_per_month:.1f} collections per month",
                            "recommendation": "REDUCE from {:.0f}x to 2-3x per week - review actual needs".format(coll_per_month),
                            "potential_savings_annual": potential_savings,
                            "timeframe": "1-3 months"
                        })
                except:
                    pass
    
    return opportunities


def calculate_electricity_rate_savings(invoice, rate_type, current_rate, target_rate):
    """Calculate potential annual savings from rate reduction"""
    usage_key = {
        "peak": "peak_usage_kwh",
        "off_peak": "off_peak_usage_kwh",
        "shoulder": "shoulder_usage_kwh"
    }.get(rate_type, "total_usage_kwh")
    
    usage = safe_float(invoice.get(usage_key, 0))
    days = get_billing_period_days(invoice) or 30
    monthly_usage = usage * (30 / days) if days > 0 else usage
    annual_usage = monthly_usage * 12
    
    # Convert rates to $/kWh
    current_rate_dollars = current_rate / 100
    target_rate_dollars = target_rate / 100
    
    savings_per_kwh = current_rate_dollars - target_rate_dollars
    annual_savings = annual_usage * savings_per_kwh
    
    return max(0, annual_savings)


def calculate_potential_savings(opportunities, grouped_data, totals):
    """Calculate total potential annual savings"""
    total_savings = {
        "conservative": 0,  # Lower end estimates
        "moderate": 0,  # Mid-range estimates
        "optimistic": 0  # Higher end estimates
    }
    
    for opp in opportunities:
        savings = opp.get("potential_savings_annual", 0)
        if savings > 0:
            severity = opp.get("severity", "medium")
            if severity == "high" or severity == "critical":
                total_savings["conservative"] += savings * 0.7
                total_savings["moderate"] += savings * 0.85
                total_savings["optimistic"] += savings
            else:
                total_savings["conservative"] += savings * 0.5
                total_savings["moderate"] += savings * 0.75
                total_savings["optimistic"] += savings
    
    return total_savings


def create_analysis_sheet(wb, business_name, analysis_results):
    """Create Base 1 Analysis sheet with benchmarking and savings opportunities"""
    sheet = wb.create_sheet("Base 1 Analysis", 7)
    
    # Title
    sheet.merge_cells('A1:F1')
    cell = sheet.cell(1, 1, f"{business_name.upper()} - BASE 1 REVIEW ANALYSIS")
    cell.font = FONT_TITLE
    cell.fill = FILL_HEADER
    cell.alignment = ALIGN_CENTER
    
    row = 3
    
    # Benchmarking Results
    sheet.cell(row, 1, "BENCHMARKING RESULTS").font = FONT_SECTION
    row += 1
    sheet.cell(row, 1, "Comparison of current rates against market benchmarks (NSW/Victoria)")
    row += 2
    
    # Headers
    headers = ["Category", "Issue Type", "Flag", "Current Rate/Cost", "Market Benchmark", "Potential Annual Savings"]
    sheet.append(headers)
    
    # Format headers
    for col in range(1, len(headers) + 1):
        cell = sheet.cell(row, col)
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.alignment = ALIGN_CENTER
    
    # Column widths
    widths = [20, 25, 12, 25, 25, 25]
    for i, width in enumerate(widths, 1):
        sheet.column_dimensions[chr(64 + i)].width = width
    
    row += 1
    
    # Add opportunities
    for opp in analysis_results["opportunities"]:
        sheet.append([
            opp.get("category", ""),
            opp.get("type", ""),
            opp.get("flag", ""),
            opp.get("current_rate", "") or opp.get("current_cost", "") or opp.get("current_situation", ""),
            opp.get("market_benchmark", ""),
            format_currency(opp.get("potential_savings_annual", 0))
        ])
        row += 1
    
    # Total Potential Savings
    row += 2
    savings = analysis_results["total_potential_savings"]
    sheet.cell(row, 1, "TOTAL POTENTIAL ANNUAL SAVINGS").font = FONT_TOTAL
    row += 1
    sheet.cell(row, 1, "Conservative Estimate:").font = FONT_SECTION
    sheet.cell(row, 2, format_currency(savings["conservative"])).font = FONT_TOTAL
    sheet.cell(row, 2).fill = FILL_YELLOW
    row += 1
    sheet.cell(row, 1, "Moderate Estimate:").font = FONT_SECTION
    sheet.cell(row, 2, format_currency(savings["moderate"])).font = FONT_TOTAL
    sheet.cell(row, 2).fill = FILL_YELLOW
    row += 1
    sheet.cell(row, 1, "Optimistic Estimate:").font = FONT_SECTION
    sheet.cell(row, 2, format_currency(savings["optimistic"])).font = FONT_TOTAL
    sheet.cell(row, 2).fill = FILL_YELLOW_TOTAL
    
    # Critical Issues
    if analysis_results["critical_issues"]:
        row += 2
        sheet.cell(row, 1, "🔴 CRITICAL ISSUES - IMMEDIATE ATTENTION REQUIRED").font = FONT_SECTION
        row += 1
        for issue in analysis_results["critical_issues"]:
            sheet.cell(row, 1, f"• {issue.get('issue', '')}")
            sheet.cell(row, 2, f"Potential saving: {format_currency(issue.get('potential_savings_annual', 0))}/year")
            row += 1
    
    # Immediate Actions
    if analysis_results["immediate_actions"]:
        row += 2
        sheet.cell(row, 1, "IMMEDIATE ACTIONS (0-30 days)").font = FONT_SECTION
        row += 1
        for action in analysis_results["immediate_actions"]:
            sheet.cell(row, 1, f"• {action.get('recommendation', '')}")
            sheet.cell(row, 2, f"Potential saving: {format_currency(action.get('potential_savings_annual', 0))}/year")
            row += 1
    
    # Short-term Actions
    if analysis_results["short_term_actions"]:
        row += 2
        sheet.cell(row, 1, "SHORT-TERM ACTIONS (1-3 months)").font = FONT_SECTION
        row += 1
        for action in analysis_results["short_term_actions"]:
            sheet.cell(row, 1, f"• {action.get('recommendation', '')}")
            sheet.cell(row, 2, f"Potential saving: {format_currency(action.get('potential_savings_annual', 0))}/year")
            row += 1
    
    # Next Steps for Base 2 Review
    row += 2
    sheet.cell(row, 1, "NEXT STEPS FOR BASE 2 REVIEW").font = FONT_SECTION
    row += 1
    next_steps = [
        "Obtain detailed rate schedules from retailers",
        "Conduct site audit for waste optimization",
        "Review demand management options",
        "Analyze usage patterns in detail",
        "Obtain competitive quotes from multiple suppliers"
    ]
    for step in next_steps:
        sheet.cell(row, 1, f"• {step}")
        row += 1


def create_meter_details_sheet(wb, grouped_data):
    """Create Meter Details sheet (Sheet 7)"""
    sheet = wb.create_sheet("Meter Details", 6)
    
    # Title
    sheet.merge_cells('A1:I1')
    cell = sheet.cell(1, 1, "ALL METER NUMBERS & READINGS")
    cell.font = FONT_TITLE
    cell.fill = FILL_HEADER
    cell.alignment = ALIGN_CENTER
    
    row = 3
    
    # Headers
    headers = [
        "Site", "Utility", "Meter/Device Number", "NMI/MIRN",
        "Last Reading Date", "Previous Read", "Current Read", "Consumption", "Units"
    ]
    sheet.append(headers)
    
    # Format headers
    for col in range(1, len(headers) + 1):
        cell = sheet.cell(row, col)
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.alignment = ALIGN_CENTER
    
    # Column widths
    widths = [25, 15, 25, 20, 20, 18, 18, 20, 10]
    for i, width in enumerate(widths, 1):
        sheet.column_dimensions[chr(64 + i)].width = width
    
    row = 4
    
    # Electricity meters
    for nmi, invoices in grouped_data["electricity"].items():
        for invoice in invoices:
            site = safe_str(invoice.get("site_address", ""))
            meter_num = safe_str(invoice.get("meter_number", "")) or nmi
            reading_date = format_date(invoice.get("invoice_date", ""))
            prev_read = safe_str(invoice.get("previous_reading", ""))
            curr_read = safe_str(invoice.get("current_reading", ""))
            consumption = safe_float(invoice.get("total_usage_kwh", 0))
            
            sheet.append([
                site, "Electricity", meter_num, nmi, reading_date,
                prev_read, curr_read, format_number(consumption), "kWh"
            ])
            row += 1
    
    # Gas meters
    for mrin, invoices in grouped_data["gas"].items():
        for invoice in invoices:
            site = safe_str(invoice.get("site_address", ""))
            meter_num = safe_str(invoice.get("meter_number", "")) or mrin
            reading_date = format_date(invoice.get("invoice_date", ""))
            prev_read = safe_str(invoice.get("previous_reading", ""))
            curr_read = safe_str(invoice.get("current_reading", ""))
            consumption = safe_float(invoice.get("total_usage_gj", 0))
            
            sheet.append([
                site, "Gas", meter_num, mrin, reading_date,
                prev_read, curr_read, format_number(consumption), "GJ"
            ])
            row += 1
    
    # Water meters
    for account, invoices in grouped_data["water"].items():
        for invoice in invoices:
            site = safe_str(invoice.get("site_address", ""))
            meter_num = safe_str(invoice.get("meter_number", "")) or account
            reading_date = format_date(invoice.get("invoice_date", ""))
            prev_read = safe_str(invoice.get("previous_reading", ""))
            curr_read = safe_str(invoice.get("current_reading", ""))
            consumption = safe_float(invoice.get("quantity", 0))
            
            sheet.append([
                site, "Water", meter_num, account, reading_date,
                prev_read, curr_read, format_number(consumption), "kL"
            ])
            row += 1


def generate_summary_pdf(run_id: str) -> str:
    """Generate Base1 Summary PDF"""
    runs = get_runs()
    if run_id not in runs:
        raise ValueError(f"Run {run_id} not found")
    
    run = runs[run_id]
    run_dir = BASE1_STORAGE / run_id / "outputs"
    
    pdf_path = run_dir / f"Base1_Summary_{run_id[:8]}.pdf"
    run_dir.mkdir(parents=True, exist_ok=True)
    
    # Create PDF
    story = []
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#1a1a1a'),
        spaceAfter=30,
        alignment=1  # Center
    )
    
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=14,
        textColor=colors.HexColor('#366092'),
        spaceAfter=12,
        spaceBefore=12
    )
    
    # Title
    story.append(Paragraph("Base 1 Review - Summary", title_style))
    story.append(Spacer(1, 0.3*inch))
    
    # Run Information
    story.append(Paragraph("Run Information", heading_style))
    story.append(Paragraph(f"<b>Run ID:</b> {run_id}", styles['Normal']))
    story.append(Paragraph(f"<b>Business Name:</b> {run.get('business_name', 'Extracted from invoices')}", styles['Normal']))
    if run.get('email'):
        story.append(Paragraph(f"<b>Email:</b> {run['email']}", styles['Normal']))
    if run.get('state'):
        story.append(Paragraph(f"<b>State:</b> {run['state']}", styles['Normal']))
    story.append(Paragraph(f"<b>Created Date:</b> {run['created_at']}", styles['Normal']))
    story.append(Spacer(1, 0.2*inch))
    
    # Documents List
    story.append(Paragraph("Uploaded Documents", heading_style))
    
    if run['documents']:
        doc_data = [["Filename", "Upload Date", "Size", "Pages"]]
        for document in run['documents']:
            upload_date = document.get('created_at', 'N/A')
            if upload_date != 'N/A':
                try:
                    dt = datetime.fromisoformat(upload_date)
                    upload_date = dt.strftime("%Y-%m-%d %H:%M")
                except:
                    pass
            
            size_mb = document.get('size_bytes', 0) / (1024 * 1024)
            doc_data.append([
                document.get('filename', 'N/A'),
                upload_date,
                f"{size_mb:.2f} MB",
                str(document.get('page_count', 0))
            ])
        
        doc_table = Table(doc_data)
        doc_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
        ]))
        story.append(doc_table)
    else:
        story.append(Paragraph("No documents uploaded.", styles['Normal']))
    
    story.append(Spacer(1, 0.3*inch))
    
    # Disclaimer
    story.append(Paragraph("Important Disclaimer", heading_style))
    disclaimer_text = (
        "Base 1 is indicative and invoice-only. This is a preliminary extraction "
        "based on uploaded documents. Full data extraction and analysis requires "
        "Base 2 review, which requires a Letter of Authority (LOA) and service fee."
    )
    story.append(Paragraph(disclaimer_text, styles['Normal']))
    story.append(Spacer(1, 0.2*inch))
    
    # Build PDF
    pdf_doc = SimpleDocTemplate(str(pdf_path), pagesize=A4)
    pdf_doc.build(story)
    
    logger.info(f"Generated summary PDF: {pdf_path}")
    return str(pdf_path)


def get_run(run_id: str) -> Dict:
    """Get run details"""
    runs = get_runs()
    if run_id not in runs:
        raise ValueError(f"Run {run_id} not found")
    return runs[run_id]

