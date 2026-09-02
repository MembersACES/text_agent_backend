"""Fill the DMA contract-details workbook and file it next to the engagement form."""

from __future__ import annotations

import io
import json
import logging
import re
from datetime import datetime
from pathlib import Path
from typing import Any, Optional

from openpyxl import load_workbook

from tools.drive_file_metadata import extract_drive_file_id
from tools.member_folder_drive import MemberFolderDriveError, upload_bytes_to_folder
from tools.one_month_savings import (
    extract_folder_id_from_url,
    get_drive_service,
)

logger = logging.getLogger(__name__)

TEMPLATE_PATH = Path(__file__).resolve().parent.parent / "templates" / "dma_contract_details.xlsx"
XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

HEADER_ALIASES: dict[str, tuple[str, ...]] = {
    "row_number": ("#", "no", "number"),
    "nmi": ("nmi",),
    "business": ("business", "business name", "company"),
    "abn": ("abn",),
    "postal_address": ("main address", "postal address", "mainaddress"),
    "site_address": ("site address", "siteaddress"),
    "frmp": ("frmp", "retailer"),
    "contact": ("contact", "contact name"),
    "position": ("postion", "position"),
    "telephone": ("contact number", "telephone", "phone"),
    "email": ("email",),
    "meter": ("meter", "dma price", "metering"),
    "vas": ("vas",),
    "start_date": ("dma start date", "start date"),
    "end_date": ("dma end date", "end date"),
}

_SPACE_RE = re.compile(r"\s+")


def _norm_header(value: Any) -> str:
    text = str(value or "").strip().lower().replace("&", "and")
    return _SPACE_RE.sub(" ", text)


def _format_date(value: str) -> str:
    raw = (value or "").strip()
    if not raw:
        return ""
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(raw[:10], fmt).strftime("%d/%m/%Y")
        except ValueError:
            continue
    return raw


def _maybe_number(value: Any) -> Any:
    if isinstance(value, bool) or value is None:
        return value
    if isinstance(value, (int, float)):
        return value
    text = str(value).strip().replace(",", "").replace("$", "")
    if not text:
        return ""
    try:
        number = float(text)
    except ValueError:
        return value
    return int(number) if number.is_integer() else number


def _optional_int(value: Any) -> Optional[int]:
    if value is None or value == "":
        return None
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def client_folder_url_from_crm(db: Any, offer_id: Any, client_id: Any) -> str:
    if db is None:
        return ""
    from models import Client, Offer

    client = None
    resolved_client_id = _optional_int(client_id)
    if resolved_client_id is not None:
        client = db.query(Client).filter(Client.id == resolved_client_id).first()
    if client is None:
        resolved_offer_id = _optional_int(offer_id)
        if resolved_offer_id is not None:
            offer = db.query(Offer).filter(Offer.id == resolved_offer_id).first()
            if offer and offer.client_id:
                client = db.query(Client).filter(Client.id == offer.client_id).first()
    return str(client.gdrive_folder_url or "") if client else ""


def _safe_filename_part(value: str) -> str:
    cleaned = re.sub(r"[^\w\s.-]", "", (value or "").strip())
    cleaned = _SPACE_RE.sub(" ", cleaned).strip(" .")
    return cleaned[:80] or "member"


def build_filename(business_name: str, nmi: str) -> str:
    return f"DMA contract details - {_safe_filename_part(business_name)} - {_safe_filename_part(nmi)}.xlsx"


def row_values(payload: dict[str, Any]) -> dict[str, Any]:
    meter = payload.get("meter") or payload.get("dma_price") or ""
    vas = payload.get("vas") or payload.get("vas_price") or ""
    return {
        "row_number": payload.get("row_number") or 1,
        "nmi": (payload.get("nmi") or "").strip(),
        "business": (payload.get("business") or payload.get("business_name") or "").strip(),
        "abn": (payload.get("abn") or "").strip(),
        "postal_address": (payload.get("postal_address") or payload.get("main_address") or "").strip(),
        "site_address": (payload.get("site_address") or "").strip(),
        "frmp": (payload.get("frmp") or payload.get("retailer") or "").strip(),
        "contact": (payload.get("contact") or payload.get("contact_name") or "").strip(),
        "position": (payload.get("position") or "").strip(),
        "telephone": (payload.get("telephone") or payload.get("contact_number") or "").strip(),
        "email": (payload.get("email") or "").strip(),
        "meter": _maybe_number(meter),
        "vas": _maybe_number(vas),
        "start_date": _format_date(str(payload.get("start_date") or payload.get("dma_start_date") or "")),
        "end_date": _format_date(str(payload.get("end_date") or payload.get("dma_end_date") or "")),
    }


def fill_workbook_bytes(payload: dict[str, Any], template_path: Optional[Path] = None) -> bytes:
    path = template_path or TEMPLATE_PATH
    if not path.is_file():
        raise FileNotFoundError(f"DMA contract details template missing: {path}")
    wb = load_workbook(path)
    ws = wb.active
    values = row_values(payload)
    key_by_col: dict[int, str] = {}
    for cell in ws[1]:
        header = _norm_header(cell.value)
        if not header:
            continue
        for key, aliases in HEADER_ALIASES.items():
            if header in aliases:
                key_by_col[cell.column] = key
                break
    if "nmi" not in key_by_col.values():
        raise ValueError("DMA contract details template is missing an NMI column.")
    for col, key in key_by_col.items():
        ws.cell(row=2, column=col, value=values.get(key, ""))
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def resolve_target_folder_id(
    engagement_form_link: str = "",
    client_folder_url: str = "",
    drive_service: Any = None,
) -> Optional[str]:
    drive = drive_service or get_drive_service()
    form_id = extract_drive_file_id(engagement_form_link)
    if form_id and drive:
        try:
            meta = drive.files().get(
                fileId=form_id,
                fields="id, mimeType, parents",
                supportsAllDrives=True,
            ).execute()
            if meta.get("mimeType") == "application/vnd.google-apps.folder":
                return form_id
            parents = [str(p) for p in (meta.get("parents") or []) if p]
            if parents:
                return parents[0]
        except Exception:
            logger.warning("Could not resolve parent folder from engagement form %s", form_id, exc_info=True)
    return extract_folder_id_from_url(client_folder_url)


def file_dma_contract_details(
    payload: dict[str, Any],
    *,
    drive_service: Any = None,
    db: Any = None,
    user_access_token: Optional[str] = None,
) -> dict[str, Any]:
    nmi = str(payload.get("nmi") or "").strip()
    business_name = str(payload.get("business") or payload.get("business_name") or "").strip()
    if not nmi:
        return {"status": "error", "message": "nmi is required."}

    drive = drive_service or get_drive_service()
    if not drive:
        return {"status": "error", "message": "Google Drive is not configured."}

    client_folder_url = str(payload.get("client_folder_url") or "").strip()
    if not client_folder_url:
        client_folder_url = client_folder_url_from_crm(
            db, payload.get("offer_id"), payload.get("client_id")
        )

    folder_id = resolve_target_folder_id(
        engagement_form_link=str(payload.get("engagement_form_link") or ""),
        client_folder_url=client_folder_url,
        drive_service=drive,
    )
    if not folder_id:
        return {
            "status": "error",
            "message": "Could not find the engagement form folder. Pass engagement_form_link or client_folder_url.",
        }

    try:
        file_bytes = fill_workbook_bytes(payload)
    except Exception as exc:
        logger.exception("Failed to fill DMA contract details workbook")
        return {"status": "error", "message": str(exc)}

    filename = build_filename(business_name, nmi)
    token = (user_access_token or "").strip() or None
    try:
        uploaded = upload_bytes_to_folder(
            file_bytes,
            filename,
            folder_id,
            mimetype=XLSX_MIME,
            drive=drive,
            user_access_token=token,
        )
    except MemberFolderDriveError as exc:
        logger.warning("DMA contract details upload failed: %s", exc.message)
        hint = ""
        if "storage quota" in (exc.message or "").lower() or "storagequotaexceeded" in (
            exc.message or ""
        ).lower():
            hint = (
                " Re-auth Google in the dashboard so the file can be uploaded as you, "
                "or move the member folder into a Shared Drive."
            )
        return {"status": "error", "message": f"{exc.message}{hint}"}
    file_id = str(uploaded.get("id") or "").strip()
    if not file_id:
        return {"status": "error", "message": "Failed to upload DMA contract details to Drive."}

    file_link = str(uploaded.get("url") or f"https://drive.google.com/file/d/{file_id}/view")
    folder_url = f"https://drive.google.com/drive/folders/{folder_id}"
    result = {
        "status": "success",
        "file_id": file_id,
        "file_link": file_link,
        "file_name": filename,
        "folder_id": folder_id,
        "folder_url": folder_url,
    }
    extra: dict[str, Any] = {
        "dma_contract_details_file_id": file_id,
        "dma_contract_details_link": file_link,
        "dma_contract_details_file_name": filename,
    }
    ef_link = str(payload.get("engagement_form_link") or "").strip()
    ef_id = extract_drive_file_id(ef_link)
    if ef_id:
        extra["engagement_form_file_id"] = ef_id
        extra["engagement_form_link"] = ef_link
    attach_contract_details_to_offer(db, payload.get("offer_id"), extra)
    return result


def attach_contract_details_to_offer(
    db: Any,
    offer_id: Any,
    extra: dict[str, Any],
) -> None:
    resolved_offer_id = _optional_int(offer_id)
    if not resolved_offer_id or db is None:
        return
    from models import OfferActivity

    activity = (
        db.query(OfferActivity)
        .filter(OfferActivity.offer_id == resolved_offer_id)
        .filter(
            OfferActivity.activity_type.in_(
                ("dma_review_generated", "dma_email_sent", "engagement_form")
            )
        )
        .order_by(OfferActivity.created_at.desc())
        .first()
    )
    if not activity:
        return
    current: dict[str, Any] = {}
    raw = activity.metadata_
    if raw:
        try:
            current = json.loads(raw) if isinstance(raw, str) else dict(raw)
        except (TypeError, ValueError):
            current = {}
    current.update(extra)
    activity.metadata_ = json.dumps(current)
    db.add(activity)
    db.commit()
