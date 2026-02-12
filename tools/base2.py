"""
Base 2 Review - Strategy Generation
Uses Base 1 Excel and invoices to generate strategy and combined Excel
"""
import os
import json
import uuid
import logging
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Optional
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

logger = logging.getLogger(__name__)

# Storage base directory
STORAGE_BASE = Path("storage")
BASE2_STORAGE = STORAGE_BASE / "base2"
RUNS_FILE = BASE2_STORAGE / "runs.json"

# Ensure storage directories exist
BASE2_STORAGE.mkdir(parents=True, exist_ok=True)


def init_storage():
    """Ensure storage directories exist"""
    BASE2_STORAGE.mkdir(parents=True, exist_ok=True)


def get_runs() -> Dict:
    """Get all runs"""
    if not RUNS_FILE.exists():
        return {}
    try:
        with open(RUNS_FILE, 'r') as f:
            return json.load(f)
    except:
        return {}


def save_runs(runs: Dict):
    """Save runs to file"""
    RUNS_FILE.parent.mkdir(parents=True, exist_ok=True)
    with open(RUNS_FILE, 'w') as f:
        json.dump(runs, f, indent=2)


def create_run(business_name: Optional[str] = None, business_info: Optional[Dict] = None) -> str:
    """Create a new Base2 run with optional business details"""
    runs = get_runs()
    run_id = str(uuid.uuid4())
    
    runs[run_id] = {
        "run_id": run_id,
        "created_at": datetime.now().isoformat(),
        "business_name": business_name,
        "business_info": business_info or {},
        "base1_excel": None,
        "invoices": [],
        "strategy": None,
        "generated": False
    }
    
    save_runs(runs)
    
    # Create run directory
    run_dir = BASE2_STORAGE / run_id
    run_dir.mkdir(parents=True, exist_ok=True)
    (run_dir / "invoices").mkdir(exist_ok=True)
    (run_dir / "outputs").mkdir(exist_ok=True)
    
    logger.info(f"Created Base2 run: {run_id} (business: {business_name or 'N/A'})")
    return run_id


def save_base1_excel(run_id: str, filename: str, content: bytes) -> Dict:
    """Save Base 1 Excel file"""
    runs = get_runs()
    if run_id not in runs:
        raise ValueError(f"Run {run_id} not found")
    
    run_dir = BASE2_STORAGE / run_id
    saved_filename = f"base1_{uuid.uuid4().hex[:8]}_{filename}"
    file_path = run_dir / saved_filename
    
    with open(file_path, 'wb') as f:
        f.write(content)
    
    runs[run_id]["base1_excel"] = {
        "filename": filename,
        "saved_filename": saved_filename,
        "size_bytes": len(content),
        "uploaded_at": datetime.now().isoformat()
    }
    
    save_runs(runs)
    logger.info(f"Saved Base1 Excel for run {run_id}: {saved_filename}")
    
    return runs[run_id]["base1_excel"]


def save_invoice(run_id: str, filename: str, content: bytes) -> Dict:
    """Save invoice PDF"""
    runs = get_runs()
    if run_id not in runs:
        raise ValueError(f"Run {run_id} not found")
    
    run_dir = BASE2_STORAGE / run_id / "invoices"
    saved_filename = f"{uuid.uuid4().hex[:8]}_{filename}"
    file_path = run_dir / saved_filename
    
    with open(file_path, 'wb') as f:
        f.write(content)
    
    doc_id = str(uuid.uuid4())
    doc_record = {
        "doc_id": doc_id,
        "filename": filename,
        "saved_filename": saved_filename,
        "size_bytes": len(content),
        "created_at": datetime.now().isoformat()
    }
    
    runs[run_id]["invoices"].append(doc_record)
    save_runs(runs)
    
    logger.info(f"Saved invoice for run {run_id}: {saved_filename}")
    return doc_record


def get_run(run_id: str) -> Dict:
    """Get run details"""
    runs = get_runs()
    if run_id not in runs:
        raise ValueError(f"Run {run_id} not found")
    return runs[run_id]


def generate_strategy_excel(run_id: str) -> str:
    """Generate Base 2 strategy Excel (1-page combined)"""
    runs = get_runs()
    if run_id not in runs:
        raise ValueError(f"Run {run_id} not found")
    
    run = runs[run_id]
    run_dir = BASE2_STORAGE / run_id / "outputs"
    run_dir.mkdir(parents=True, exist_ok=True)
    
    # Load Base 1 Excel
    if not run.get("base1_excel"):
        raise ValueError("Base 1 Excel not uploaded")
    
    base1_excel_path = BASE2_STORAGE / run_id / run["base1_excel"]["saved_filename"]
    if not base1_excel_path.exists():
        raise ValueError("Base 1 Excel file not found")
    
    # Load the Base 1 workbook
    base1_wb = load_workbook(str(base1_excel_path))
    
    # Create new Base 2 workbook
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    
    # Create strategy sheet
    strategy_sheet = wb.create_sheet("Base 2 Strategy", 0)
    
    # Get business name from run data, Base 1, or default
    business_name = run.get("business_name") or "Business Name"
    if business_name == "Business Name":
        # Try to get from Base 1 Overview sheet
        if "Overview" in base1_wb.sheetnames:
            overview_sheet = base1_wb["Overview"]
            # Look for "BASE 1 REVIEW - [NAME]" in first row
            if overview_sheet["A1"].value:
                name_str = str(overview_sheet["A1"].value)
                if "BASE 1 REVIEW -" in name_str:
                    business_name = name_str.replace("BASE 1 REVIEW -", "").strip()
        # Fallback to old Summary sheet
        elif "Summary Base1 Review" in base1_wb.sheetnames:
            summary_sheet = base1_wb["Summary Base1 Review"]
            if summary_sheet["A1"].value:
                business_name = str(summary_sheet["A1"].value)
    
    # Header
    strategy_sheet.append([business_name, "", "", "", ""])
    strategy_sheet.append(["Base 2 Review - Strategy", "", "", "", ""])
    strategy_sheet.append([])
    
    # Extract key data from Base 1 Summary sheet
    if "Summary Base1 Review" in base1_wb.sheetnames:
        summary_sheet = base1_wb["Summary Base1 Review"]
        
        # Find current totals
        current_total = None
        estimated_total = None
        savings = None
        
        for row in summary_sheet.iter_rows():
            row_values = [cell.value for cell in row]
            if any("Grand Total Annual" in str(v) for v in row_values if v):
                # Try to extract the total
                for cell in row:
                    if cell.value and "$" in str(cell.value):
                        current_total = str(cell.value)
            if any("All NMI Total Annual" in str(v) for v in row_values if v):
                for cell in row:
                    if cell.value and "$" in str(cell.value) and "Estimated" in str(row_values[0] or ""):
                        estimated_total = str(cell.value)
            if any("Annual estimated savings" in str(v) for v in row_values if v):
                for cell in row:
                    if cell.value and "$" in str(cell.value):
                        savings = str(cell.value)
        
        # Strategy Summary
        strategy_sheet.append(["Strategy Summary", "", "", "", ""])
        strategy_sheet.append([])
        
        if current_total:
            strategy_sheet.append(["Current Annual Cost", "", "", "", current_total])
        if estimated_total:
            strategy_sheet.append(["Estimated Annual Cost", "", "", "", estimated_total])
        if savings:
            strategy_sheet.append(["Potential Annual Savings", "", "", "", savings])
        
        strategy_sheet.append([])
        
        # Key Recommendations
        strategy_sheet.append(["Key Recommendations", "", "", "", ""])
        strategy_sheet.append([])
        
        # Extract opportunities from Base 1 data
        recommendations = [
            "Review electricity rates - current rates may be above market benchmarks",
            "Consider competitive tender for better rates",
            "Optimize demand charges through load management",
            "Review metering charges - potential for cost reduction",
            "Consolidate accounts where possible for better rates"
        ]
        
        for i, rec in enumerate(recommendations, 1):
            strategy_sheet.append([f"{i}.", rec, "", "", ""])
        
        strategy_sheet.append([])
        
        # Action Items
        strategy_sheet.append(["Action Items", "", "", "", ""])
        strategy_sheet.append([])
        action_items = [
            ["1.", "Obtain Letter of Authority (LOA) from client", "", "", ""],
            ["2.", "Conduct full utility data extraction", "", "", ""],
            ["3.", "Prepare competitive tender documents", "", "", ""],
            ["4.", "Engage with multiple energy retailers", "", "", ""],
            ["5.", "Negotiate best rates and terms", "", "", ""],
            ["6.", "Implement new contracts", "", "", ""],
        ]
        
        for item in action_items:
            strategy_sheet.append(item)
        
        strategy_sheet.append([])
        
        # Next Steps
        strategy_sheet.append(["Next Steps", "", "", "", ""])
        strategy_sheet.append([])
        strategy_sheet.append(["Timeline", "Activity", "Responsible", "Status", "Notes"])
        
        timeline = [
            ["Week 1", "LOA Collection & Data Extraction", "Account Manager", "Pending", ""],
            ["Week 2-3", "Tender Preparation", "Energy Consultant", "Pending", ""],
            ["Week 4", "Tender Submission", "Energy Consultant", "Pending", ""],
            ["Week 5-6", "Review & Negotiate", "Account Manager", "Pending", ""],
            ["Week 7", "Contract Execution", "Legal & Client", "Pending", ""],
            ["Week 8+", "Implementation", "Operations", "Pending", ""],
        ]
        
        for item in timeline:
            strategy_sheet.append(item)
    
    # Style the sheet
    strategy_sheet["A1"].font = Font(bold=True, size=16)
    strategy_sheet["A2"].font = Font(bold=True, size=14)
    
    # Style section headers
    for row in strategy_sheet.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                if any(keyword in cell.value for keyword in ["Strategy Summary", "Key Recommendations", "Action Items", "Next Steps"]):
                    cell.font = Font(bold=True, size=12)
                    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    cell.font = Font(bold=True, color="FFFFFF", size=12)
    
    # Auto-adjust column widths
    for col in strategy_sheet.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        strategy_sheet.column_dimensions[col_letter].width = adjusted_width
    
    # Save workbook
    date_str = datetime.now().strftime("%Y.%m.%d")
    excel_path = run_dir / f"{business_name.replace(' ', '_')}_Base2_Strategy_{date_str}.xlsx"
    wb.save(excel_path)
    
    logger.info(f"Generated Base2 strategy Excel: {excel_path}")
    return str(excel_path)

